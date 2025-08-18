import os
import discord
import pandas as pd
import mysql.connector
from datetime import datetime
from discord.ext import commands
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font
from database import db

# Load environment variables
load_dotenv()

# Bot setup with intents
intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot = commands.Bot(command_prefix='!', intents=intents)

# Dictionary to store user data during onboarding
user_data = {}

# Function to save data to Excel
def save_identity_data(user_name, data, filename="onboarding_data.xlsx"):
    """Save identity data to a separate Excel file using user's name"""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Create a new filename for identity data
        base_filename = os.path.splitext(filename)[0]
        identity_filename = f"{base_filename}_identity.xlsx"
        
        # Load or create identity workbook
        if os.path.exists(identity_filename):
            book = load_workbook(identity_filename)
            ws = book.active
            
            # Check if headers exist, if not add them
            if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                headers = ['Full Name', 'ID Number', 'Passport', 'KRA Number', 'Last Updated']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        else:
            # Create new workbook with headers
            book = Workbook()
            ws = book.active
            ws.title = "Identity Data"
            headers = ['Full Name', 'ID Number', 'Passport', 'KRA Number', 'Last Updated']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        # Find if user exists
        user_found = False
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value).lower() == str(user_name).lower():
                # Update existing entry
                ws.cell(row=row, column=2, value=data.get('id_number', 'N/A'))
                ws.cell(row=row, column=3, value=data.get('passport', 'N/A'))
                ws.cell(row=row, column=4, value=data.get('kra', 'Not provided'))
                ws.cell(row=row, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                user_found = True
                break
        
        # If user not found, add new row
        if not user_found:
            new_row = [
                user_name,
                data.get('id_number', 'N/A'),
                data.get('passport', 'N/A'),
                data.get('kra', 'Not provided'),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
            ws.append(new_row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 30)
        
        # Save to the new file
        book.save(identity_filename)
        return True
    except Exception as e:
        print(f"Error saving identity data: {e}")
        return False

def save_to_excel(user_id, data, filename="onboarding_data.xlsx"):
    """
    Save user data to both Excel and MySQL database.
    Returns the entry code if successful, None otherwise.
    """
    import os
    import string
    import random
    from datetime import datetime
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    entry_code = None
    db_success = False
    excel_success = False
    
    try:
        
        # Function to generate a random alphanumeric code
        def generate_unique_code(length=8):
            chars = string.ascii_uppercase + string.digits  # A-Z and 0-9
            return ''.join(random.choice(chars) for _ in range(length))
        
        # Get file path if it exists
        file_path = data.get('file_path', 'None')
        
        # Generate a unique code for this entry
        entry_code = generate_unique_code()
        
        # Save to MySQL database first
        try:
            # Prepare data for database
            db_data = data.copy()
            db_data['file_path'] = file_path if file_path != 'None' else 'No file uploaded'
            
            # Save to database
            entry_code = db.save_member(user_id, db_data)
            if not entry_code:
                print("Failed to save to database")
                return None
                
            print(f"Successfully saved to database with entry code: {entry_code}")
            db_success = True
                
        except Exception as e:
            print(f"Error saving to database: {e}")
            # Continue with Excel save even if database save fails
        
        # Continue with Excel save
        try:
            # Create parent directories if they don't exist
            os.makedirs(os.path.dirname(os.path.abspath(filename)) or '.', exist_ok=True)
            
            # Initialize DataFrame for Excel
            df = pd.DataFrame()
            
            # Check if file exists and is not empty
            if os.path.exists(filename) and os.path.getsize(filename) > 0:
                try:
                    # Read existing data
                    df = pd.read_excel(filename)
                    
                    # Ensure required columns exist
                    required_columns = [
                        'Entry Code', 'User ID', 'Full Name', 'Email', 
                        'Phone', 'Date of Birth', 'File Path', 
                        'Registration Date', 'Status'
                    ]
                    
                    # Add missing columns
                    for col in required_columns:
                        if col not in df.columns:
                            df[col] = ''
                    
                except Exception as e:
                    print(f"Error reading existing file: {e}")
                    df = pd.DataFrame(columns=required_columns)
        except Exception as e:
            print(f"Error setting up Excel file: {e}")
            return None
            
            # Create a DataFrame with the user's data (exclude sensitive data from main sheet)
            new_data = {
                'Entry Code': entry_code,  # Use the same entry code from database
                'User ID': str(user_id),
                'Full Name': data.get('full_name', '').title(),
                'Email': data.get('email', '').lower(),
                'Phone': data.get('phone', ''),
                'Date of Birth': data.get('dob', ''),
                'File Path': file_path if file_path != 'None' else 'No file uploaded',
                'Registration Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Status': 'Active'
            }
            
            # Save sensitive data to separate sheet first
            user_name = data.get('full_name', '').title()
            save_identity_data(user_name, data, filename)
            
            # Always append new entry - never update existing ones
            if os.path.exists(filename):
                try:
                    # Read existing data
                    df = pd.read_excel(filename)
                    # Ensure Entry Code column exists
                    if 'Entry Code' not in df.columns:
                        df['Entry Code'] = ''
                    # Generate new unique code if it already exists
                    while True:
                        existing_codes = set(df['Entry Code'].dropna())  # Get all existing codes
                        if entry_code not in existing_codes:
                            break
                        entry_code = generate_unique_code()
                        new_data['Entry Code'] = entry_code
                    # Append new data
                    df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)
                except Exception as e:
                    print(f"Error reading existing file: {e}")
                    df = pd.DataFrame([new_data])
            else:
                # Create new DataFrame with the new entry
                df = pd.DataFrame([new_data])
                
            # Create Excel writer object with error handling
            try:
                writer = pd.ExcelWriter(filename, engine='openpyxl')
                excel_success = True
            except Exception as e:
                print(f"Error creating Excel writer: {e}")
                if db_success:
                    print("Data was saved to database but not to Excel")
                    return entry_code
                return None
                
                # Final check before saving
                if not excel_success and not db_success:
                    print("Failed to save data to both database and Excel")
                    return None
                
                # Ensure all required columns exist
                required_columns = [
                    'Entry Code',
                    'User ID',
                    'Full Name',
                    'Email',
                    'Phone',
                    'Date of Birth',
                    'Registration Date',
                    'Status'
                ]
            
                # Add any missing columns with empty values
                for col in required_columns:
                    if col not in df.columns:
                        df[col] = ''
                
                # Reorder columns to match required order
                df = df[required_columns]
                
                # Convert date strings to datetime objects for proper Excel formatting
                if 'Registration Date' in df.columns:
                    df['Registration Date'] = pd.to_datetime(df['Registration Date'])
                
                # Define styles
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
                
                # Write DataFrame to Excel
                df.to_excel(writer, index=False, sheet_name='Onboarding Data')
        
        # Get the worksheet
        if 'Onboarding Data' in writer.sheets:
            worksheet = writer.sheets['Onboarding Data']
        else:
            worksheet = writer.book['Onboarding Data']
        
        # Apply styles to header row
        if worksheet.max_row > 0:  # Check if there are any rows
            for cell in worksheet[1]:  # First row is header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths and apply styles to data rows
        if worksheet.max_row > 1:  # Check if there are data rows
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find the maximum length of content in the column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value or ''))
                    except:
                        pass
                
                # Set column width (with some extra space)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 30)  # Min width 10, max 30
                
                # Apply styles to all cells
                for cell in column:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Auto-filter for the header row
        if worksheet.max_row > 1:  # Only add filter if there are data rows
            worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        
        # Save the workbook
        writer.close()
        return True
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

# Function to ask the next question
async def validate_id_number(id_number):
    """Validate ID number (5-9 digits)"""
    return id_number.isdigit() and 5 <= len(id_number) <= 9

def validate_passport(passport):
    """Validate passport number (2 letters followed by 5 digits)"""
    if len(passport) != 7:
        return False
    return (passport[:2].isalpha() and 
            passport[2:].isdigit() and 
            len(passport[2:]) == 5)

def validate_kra(kra):
    """Validate KRA number (alphanumeric, 11 characters)"""
    return len(kra) == 11 and kra.isalnum()

async def ask_next_question(ctx, user_id):
    questions = [
        ("full_name", "What is your full name?"),
        ("email", "What is your email address?"),
        ("phone", "What is your phone number?"),
        ("dob", "What is your date of birth? (DD/MM/YYYY)"),
        ("id_type", "Do you have a National ID? (yes/no) - If no, you'll be asked for passport number"),
        ("kra_prompt", "Would you like to provide your KRA PIN? (yes/no)")
    ]
    
    # Find the first question that hasn't been answered yet
    for field, question in questions:
        if field not in user_data[user_id]:
            await ctx.send(f"{question}")
            return field
    
    # Handle ID/Passport flow - Only check if we're not already in the middle of handling an ID or passport
    if 'id_type' in user_data[user_id] and 'id_number' not in user_data[user_id] and 'passport' not in user_data[user_id]:
        response = user_data[user_id]['id_type'].lower()
        if response in ['yes', 'y'] and 'awaiting_input' not in user_data[user_id]:
            return 'id_number'
        elif response in ['no', 'n'] and 'awaiting_input' not in user_data[user_id]:
            return 'passport'
        elif 'awaiting_input' not in user_data[user_id]:  # Only reset if we're not in the middle of something
            del user_data[user_id]['id_type']
            return 'id_type'
    
    # Handle KRA prompt
    if 'kra_prompt' in user_data[user_id] and 'kra' not in user_data[user_id]:
        response = user_data[user_id]['kra_prompt'].lower()
        if response in ['yes', 'y']:
            return 'kra'
        elif response in ['no', 'n']:
            user_data[user_id]['kra'] = 'Not provided'
        else:
            del user_data[user_id]['kra_prompt']  # Reset if invalid response
            return 'kra_prompt'
    
    # Ask for KRA number after ID/Passport
    if ('id_number' in user_data[user_id] or 'passport' in user_data[user_id]) and 'kra' not in user_data[user_id]:
        return 'kra'
    
    # If all questions have been answered, ask about file upload
    if 'file_uploaded' not in user_data[user_id]:
        await ctx.send("Would you like to upload a file? (yes/no)")
        return 'ask_file_upload'
    
    return None

async def handle_file_upload(message, user_id):
    """Handle file upload from user"""
    if not message.attachments:
        await message.channel.send("No file attached. Please try again or type 'skip' to continue without uploading.")
        return False
    
    attachment = message.attachments[0]
    
    # Create uploads directory if it doesn't exist
    upload_dir = os.path.join('user_uploads', str(user_id))
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate a unique filename
    file_extension = os.path.splitext(attachment.filename)[1]
    unique_filename = f"{int(time.time())}_{attachment.filename}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Save the file
    try:
        await attachment.save(file_path)
        user_data[user_id]['file_path'] = file_path
        user_data[user_id]['file_uploaded'] = True
        await message.channel.send(f"✅ File uploaded successfully! ({attachment.filename})")
        return True
    except Exception as e:
        print(f"Error saving file: {e}")
        await message.channel.send("❌ Error saving file. Please try again or type 'skip' to continue.")
        return False

@bot.event
async def on_message(message):
    # Ignore messages from the bot itself
    if message.author == bot.user:
        return
    
    # Process commands first
    await bot.process_commands(message)
    
    # Handle onboarding conversation
    user_id = str(message.author.id)
    
    # If user is in the middle of onboarding and the message is not a command
    if user_id in user_data and user_data[user_id].get('awaiting_input') and not message.content.startswith('!'):
        current_field = user_data[user_id]['awaiting_input']
        
        # Handle file upload question
        if current_field == 'ask_file_upload':
            response = message.content.strip().lower()
            if response in ['yes', 'y']:
                await message.channel.send("Please upload your file by dragging and dropping it into this chat.")
                user_data[user_id]['awaiting_input'] = 'file_upload'
                return
            elif response in ['no', 'n', 'skip']:
                user_data[user_id]['file_uploaded'] = False
                user_data[user_id]['file_path'] = 'None'
                user_data[user_id]['awaiting_input'] = None
            else:
                await message.channel.send("Please answer with 'yes' or 'no'.")
                return
        # Handle actual file upload
        elif current_field == 'file_upload':
            if message.content.strip().lower() == 'skip':
                user_data[user_id]['file_uploaded'] = False
                user_data[user_id]['file_path'] = 'None'
                user_data[user_id]['awaiting_input'] = None
            else:
                success = await handle_file_upload(message, user_id)
                if success:
                    user_data[user_id]['awaiting_input'] = None
                return
        # Handle regular text inputs
        elif current_field == 'full_name':
            user_data[user_id]['full_name'] = message.content.strip()
        elif current_field == 'email':
            user_data[user_id]['email'] = message.content.strip().lower()
        elif current_field == 'phone':
            user_data[user_id]['phone'] = message.content.strip()
        elif current_field == 'id_type':
            response = message.content.strip().lower()
            if response in ['yes', 'y']:
                user_data[user_id]['id_type'] = response
                await message.channel.send("Please enter your National ID number (5-9 digits):")
                user_data[user_id]['awaiting_input'] = 'id_number'  # Set next field
                return
            elif response in ['no', 'n']:
                user_data[user_id]['id_type'] = response
                await message.channel.send("Please enter your passport number (2 letters followed by 5 digits, e.g., AB12345):")
                user_data[user_id]['awaiting_input'] = 'passport'  # Set next field
                return
            else:
                await message.channel.send("❌ Please answer with 'yes' or 'no'.")
                return
        elif current_field == 'id_number':
            id_num = message.content.strip()
            if validate_id_number(id_num):
                user_data[user_id]['id_number'] = id_num
                user_data[user_id]['awaiting_input'] = 'kra_prompt'  # Move to KRA prompt next
            else:
                await message.channel.send("❌ Invalid ID number. Please enter 5-9 digits.")
                return
        elif current_field == 'passport':
            passport = message.content.strip().upper()
            if validate_passport(passport):
                user_data[user_id]['passport'] = passport
                user_data[user_id]['awaiting_input'] = 'kra_prompt'  # Move to KRA prompt next
            else:
                await message.channel.send("❌ Invalid passport number. Format: 2 letters followed by 5 digits (e.g., AB12345).")
                return
        elif current_field == 'kra':
            kra = message.content.strip().upper()
            if validate_kra(kra):
                user_data[user_id]['kra'] = kra
                user_data[user_id]['awaiting_input'] = 'ask_file_upload'  # Move to file upload next
            else:
                await message.channel.send("❌ Invalid KRA PIN. Must be 11 alphanumeric characters. Please try again or type 'skip' to continue without providing KRA PIN.")
                return
        elif current_field == 'kra_prompt':
            response = message.content.strip().lower()
            if response in ['yes', 'y']:
                user_data[user_id]['kra_prompt'] = response
                user_data[user_id]['awaiting_input'] = 'kra'  # Ask for KRA number
                await message.channel.send("Please enter your KRA PIN (11 alphanumeric characters):")
                return
            elif response in ['no', 'n']:
                user_data[user_id]['kra_prompt'] = response
                user_data[user_id]['kra'] = 'Not provided'
                user_data[user_id]['awaiting_input'] = 'ask_file_upload'  # Skip to file upload
            else:
                await message.channel.send("❌ Please answer with 'yes' or 'no'.")
                return
        elif current_field == 'dob':
            user_data[user_id]['dob'] = message.content.strip()
        
        # Ask the next question or complete onboarding
        next_field = await ask_next_question(message.channel, user_id)
        
        if next_field:
            user_data[user_id]['awaiting_input'] = next_field
        else:
            # Save data to Excel
            if save_to_excel(user_id, user_data[user_id]):
                await message.channel.send("✅ Thank you for completing the onboarding process! Your information has been saved.")
            else:
                await message.channel.send("❌ There was an error saving your information. Please try again later.")
            # Clean up
            del user_data[user_id]

@bot.event
async def on_ready():
    print(f'We have logged in as {bot.user}')
    print('Bot is ready to receive commands!')

@bot.command(name='start')
async def start_onboarding(ctx):
    """Starts the onboarding process"""
    user_id = str(ctx.author.id)
    
    # Initialize user data
    user_data[user_id] = {}
    
    await ctx.send("Welcome to AAR Insurance! Let's get started with your onboarding.")
    
    # Ask the first question
    first_field = await ask_next_question(ctx, user_id)
    if first_field:
        user_data[user_id]['awaiting_input'] = first_field

@bot.command(name='helpme')
async def help_command(ctx):
    """Displays help information"""
    help_text = """
    **AAR Insurance Bot Commands:**
    - `!start` - Begin the onboarding process
    - `!status [entry_code] [activate|deactivate]` - View or update member status
    - `!helpme` - Show this help message
    
    **Status Management:**
    - `!status` - View all members and their statuses
    - `!status [entry_code] activate` - Activate a member
    - `!status [entry_code] deactivate` - Deactivate a member
    
    **Onboarding Process:**
    The bot will guide you through a series of questions to collect your information.
    """
    await ctx.send(help_text)

async def update_member_status(entry_code, new_status, filename="onboarding_data.xlsx"):
    """
    Update a member's status in both Excel and MySQL database.
    Returns True if successful in at least one storage, False otherwise.
    """
    success = False
    
    # Update status in MySQL
    try:
        db_updated = db.update_member_status(entry_code, new_status)
        if db_updated:
            print(f"Status for entry {entry_code} updated to {new_status} in database")
            success = True
        else:
            print(f"Failed to update status in database for entry {entry_code}")
    except Exception as e:
        print(f"Error updating status in database: {e}")
    
    # Update status in Excel
    excel_updated = False
    try:
        if not os.path.exists(filename):
            print(f"Excel file {filename} does not exist, skipping Excel update")
            return success
            
        # Read the Excel file
        df = pd.read_excel(filename)
        
        # Check if the entry code exists
        if 'Entry Code' not in df.columns:
            print("Error: 'Entry Code' column not found in the Excel file")
            return success
            
        # Find the row with the matching entry code
        mask = df['Entry Code'].astype(str) == str(entry_code)
        if not any(mask):
            print(f"Warning: Entry code {entry_code} not found in Excel, but updated in database")
            return success
            
        # Update the status
        df.loc[mask, 'Status'] = new_status
        
        # Save back to Excel with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Onboarding Data')
            worksheet = writer.sheets['Onboarding Data']
            
            # Apply formatting
            header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            header_font = Font(bold=True)
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                worksheet.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 30)
        
            # Apply formatting
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
        
            print(f"Status for entry {entry_code} updated to {new_status} in Excel")
            excel_updated = True
            
            member_name = df.loc[mask, 'Full Name'].iloc[0]
            action = "activated" if new_status.lower() == 'active' else "deactivated"
            return True, f"✅ You have successfully {action} {member_name}"
            
    except Exception as e:
        print(f"Error updating member status in Excel: {e}")
        return success or excel_updated, f"❌ Error updating status: {str(e)}"
        
    except Exception as e:
        return False, f"Error updating status: {str(e)}"


def get_all_members(filename="onboarding_data.xlsx"):
    """Get all members with their details"""
    try:
        if not os.path.exists(filename):
            return []
            
        df = pd.read_excel(filename)
        if df.empty:
            return []
            
        # Convert relevant columns to string and handle NaN values
        df = df.where(pd.notnull(df), '')
        return df.to_dict('records')
        
    except Exception as e:
        print(f"Error reading members: {e}")
        return []


def initialize_excel(filename="onboarding_data.xlsx"):
    """Initialize the Excel file with proper columns if it doesn't exist"""
    required_columns = [
        'Entry Code',
        'User ID',
        'Full Name',
        'Email',
        'Phone',
        'Date of Birth',
        'Registration Date',
        'Status'
    ]
    
    if not os.path.exists(filename):
        df = pd.DataFrame(columns=required_columns)
        df.to_excel(filename, index=False, sheet_name='Onboarding Data')
        print(f"Created new Excel file: {filename}")
    else:
        # Ensure all required columns exist in existing file
        try:
            df = pd.read_excel(filename)
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''
            df = df[required_columns]  # Reorder columns
            df.to_excel(filename, index=False, sheet_name='Onboarding Data')
            print(f"Updated existing Excel file with new columns: {filename}")
        except Exception as e:
            print(f"Error updating Excel file: {e}")
            print(f"Using existing Excel file as-is: {filename}")

# Run the bot
@bot.command(name='status')
async def status_command(ctx, entry_code: str = None, action: str = None):
    """View or update member status with database integration"""
    try:
        if entry_code and action:
            # Update status in database
            if action.lower() not in ['activate', 'deactivate']:
                await ctx.send("❌ Invalid action. Use 'activate' or 'deactivate'")
                return
                
            new_status = 'Active' if action.lower() == 'activate' else 'Inactive'
            success = db.update_member_status(entry_code, new_status)
            
            if success:
                await ctx.send(f"✅ Successfully {action}d member with entry code: {entry_code}")
            else:
                await ctx.send(f"❌ Failed to update status for entry code: {entry_code}")
                
        else:
            # Get all members from database
            try:
                members = db.get_all_members()
                if not members:
                    await ctx.send("No members found in the database. Use `!start` to register.")
                    return
                
                # Create a formatted message
                message = "**📊 Member Status (From Database)**\n\n"
                
                # Add database stats
                total_members = len(members)
                active_members = sum(1 for m in members if m[5].lower() == 'active')  # Status is at index 5
                
                message += f"👥 **Total Members:** {total_members}\n"
                message += f"🟢 **Active:** {active_members}\n"
                message += f"🔴 **Inactive:** {total_members - active_members}\n\n"
                message += "**Member List:**\n\n"
                
                # Add member details (limited to prevent message overflow)
                for member in members[:10]:  # Show first 10 members
                    status_emoji = "🟢" if member[5].lower() == 'active' else "🔴"  # Status at index 5
                    message += (
                        f"`{member[1]}` - {member[3]} "  # Entry code at 1, Full name at 3
                        f"({member[4]}) - {status_emoji} {member[5]}\n"  # Email at 4, Status at 5
                    )
                
                if len(members) > 10:
                    message += f"\n... and {len(members) - 10} more members. Use `!status [entry_code]` for details."
                
                message += "\n**Usage:**\n"
                message += "- `!status` - Show this summary\n"
                message += "- `!status [entry_code]` - Show member details\n"
                message += "- `!status [entry_code] [activate|deactivate]` - Update status"
                
                # Send the message in chunks if too long
                if len(message) > 2000:
                    chunks = [message[i:i+2000] for i in range(0, len(message), 2000)]
                    for chunk in chunks:
                        await ctx.send(chunk)
                else:
                    await ctx.send(message)
                    
            except Exception as e:
                await ctx.send(f"❌ Error fetching members from database: {str(e)}")
                
    except Exception as e:
        await ctx.send(f"❌ An error occurred: {str(e)}")


if __name__ == "__main__":
    # Initialize the Excel file with proper columns
    initialize_excel()
    
    TOKEN = os.getenv('DISCORD_TOKEN')
    if not TOKEN:
        print("Error: No Discord token found in .env file")
    else:
        bot.run(TOKEN)
